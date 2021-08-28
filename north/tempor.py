import openpyxl
import csv
import plotly.graph_objects as go
import pandas as pd
import chart_studio.plotly as py
import os

if not os.path.exists("images"):
    os.mkdir("images")

companies = {}
def getdata():
    trainsDB = openpyxl.load_workbook('StockList.xlsx')
    sheet = trainsDB['Stock List']

    currentId = str(sheet.cell(row=2, column=5).value)
    global companies
    temp = {}

    for i in range(2, sheet.max_row + 1):
        if (currentId == str(sheet.cell(row=i, column=5).value)):
            schedule = [float(sheet.cell(row=i, column=3).value),
                        float(sheet.cell(row=i, column=2).value),
                        float(sheet.cell(row=i, column=4).value),
                        float(sheet.cell(row=i, column=1).value),
                        int(sheet.cell(row=i, column=6).value)]
            temp[str(sheet.cell(row=i, column=7).value)]=schedule
        else:
            companies[currentId] = temp
            temp = {}
            currentId = str(sheet.cell(row=i, column=5).value)
            schedule = [float(sheet.cell(row=i, column=3).value),
                        float(sheet.cell(row=i, column=2).value),
                        float(sheet.cell(row=i, column=4).value),
                        float(sheet.cell(row=i, column=1).value),
                        int(sheet.cell(row=i, column=6).value)]
            temp[str(sheet.cell(row=i, column=7).value)]=schedule
    companies[currentId] = temp
    sorted (companies.keys())

getdata()

def compt(userinput,startdate,enddate):
    
    if (userinput in companies.keys()):
        data = companies[userinput]
        companykey = list(sorted(data.keys()))
        length = len(data)
        for i in range(0,companykey.index(startdate)):
            del data[companykey[i]]
        for i in range(companykey.index(enddate)+1,length):
            del data[companykey[i]]
        companykey = list(sorted(data.keys()))
        print(data)
    else:
        print('Invalid code')
    with open('result.csv','w') as output:
        output_data = csv.writer(output, delimiter = ',')
        output_data.writerow(['low','high','open','close','volume','date'])
        i=0
        for val in data.values():
            output_data.writerow([str(val[0]),str(val[1]),str(val[2]),str(val[3]),str(val[4]),companykey[i]])
            i+=1

    df = pd.read_csv("D:\\north\\result.csv")
    fig = go.Figure(data=go.Ohlc(x=df['date'],
                    open=df['open'],
                    high=df['high'],
                    low=df['low'],
                    close=df['close']))
    #fig.write_image("images/fig1.png")
    fig.show()
    url = py.plot(fig, filename='stacked-bar')
    print(url)
    fig.write_html("images/fig1.html")   

code = input('Enter the code of company')
startdate = input('Enter start date in YYYY-MM-DD format')
enddate = input('Enter end date in YYYY-MM-DD format')
compt(code, startdate, enddate)